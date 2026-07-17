from __future__ import annotations
import io
import csv
from pathlib import Path

from pypdf import PdfReader
from docx import Document
from openpyxl import load_workbook

SUPPORTED_TEXT_EXT = {".txt", ".md", ".json", ".csv"}
SUPPORTED_IMAGE_EXT = {".png", ".jpg", ".jpeg", ".webp"}
SUPPORTED_DOC_EXT = {".pdf", ".docx", ".xlsx"}

def _safe_text(text: str, limit: int = 12000) -> str:
    text = (text or "").strip()
    if len(text) > limit:
        return text[:limit] + "\n\n[... içerik kısaltıldı ...]"
    return text

def read_pdf_bytes(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    parts = []
    for page in reader.pages[:20]:
        try:
            parts.append(page.extract_text() or "")
        except Exception:
            pass
    return _safe_text("\n".join(parts))

def read_docx_bytes(data: bytes) -> str:
    doc = Document(io.BytesIO(data))
    parts = [p.text for p in doc.paragraphs if (p.text or "").strip()]
    return _safe_text("\n".join(parts))

def read_xlsx_bytes(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), data_only=True)
    lines = []
    for ws in wb.worksheets[:5]:
        lines.append(f"[Sheet] {ws.title}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
            vals = [str(v) for v in row if v is not None and str(v).strip()]
            if vals:
                lines.append(" | ".join(vals))
    return _safe_text("\n".join(lines))

def read_csv_bytes(data: bytes) -> str:
    text = data.decode("utf-8", "ignore")
    out = []
    for i, row in enumerate(csv.reader(io.StringIO(text))):
        if i >= 60:
            break
        vals = [c.strip() for c in row if c and c.strip()]
        if vals:
            out.append(" | ".join(vals))
    return _safe_text("\n".join(out))

def read_plain_bytes(data: bytes) -> str:
    return _safe_text(data.decode("utf-8", "ignore"))

def analyze_uploaded_file(filename: str, data: bytes) -> dict:
    ext = Path(filename or "").suffix.lower()

    if ext in SUPPORTED_IMAGE_EXT:
        return {
            "type": "image",
            "filename": filename,
            "content": f"[Görsel yüklendi: {filename}] Kullanıcı bu görselin analizini istiyor. Görseli açıklama, içeriğini yorumlama, nesneleri/anlamı çıkarma yönünde yanıt ver."
        }

    if ext == ".pdf":
        return {"type": "document", "filename": filename, "content": read_pdf_bytes(data)}

    if ext == ".docx":
        return {"type": "document", "filename": filename, "content": read_docx_bytes(data)}

    if ext == ".xlsx":
        return {"type": "spreadsheet", "filename": filename, "content": read_xlsx_bytes(data)}

    if ext == ".csv":
        return {"type": "spreadsheet", "filename": filename, "content": read_csv_bytes(data)}

    if ext in SUPPORTED_TEXT_EXT:
        return {"type": "text", "filename": filename, "content": read_plain_bytes(data)}

    return {
        "type": "unknown",
        "filename": filename,
        "content": f"[Desteklenmeyen veya sınırlı destekli dosya: {filename}] Kullanıcı bu dosya hakkında analiz istiyor."
    }

def build_file_context(items: list[dict]) -> str:
    blocks = []
    for item in items:
        name = item.get("filename", "dosya")
        ftype = item.get("type", "unknown")
        content = (item.get("content") or "").strip()
        blocks.append(
            f"Dosya adı: {name}\nDosya tipi: {ftype}\nİçerik:\n{content}"
        )
    return "\n\n" + ("\n\n---\n\n".join(blocks)).strip()
